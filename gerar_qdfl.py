#!/usr/bin/env python3
"""
Gera QDFL.xlsx no formato exato do modelo QDC-MODELO-127V-220V.xlsx
Chamado com: python3 gerar_qdfl.py dados.json saida.xlsx
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# ── Cores do modelo (exatas do Excel) ────────────────────────────
COR_HEADER_DARK  = "1F3864"   # azul escuro — título principal
COR_HEADER_MID   = "2F5496"   # azul médio — grupo de colunas
COR_HEADER_LIGHT = "BDD7EE"   # azul claro — subgrupos
COR_CIRC_ALT     = "EBF3FB"   # azul bem claro — linhas alternadas
COR_OK           = "C6EFCE"   # verde — OK
COR_WARN         = "FFEB9C"   # amarelo — atenção
COR_ERR          = "FFC7CE"   # vermelho — erro
COR_TOTAL        = "DDEBF7"   # azul — linha totais
COR_WHITE        = "FFFFFF"
FONTE            = "Arial"

def ft(bold=False, size=9, color="000000", name=FONTE):
    return Font(name=name, bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def aln(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style='medium', color='000000')
    st = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, bg=None, halign="center",
             valign="center", wrap=True, num_fmt=None, font_size=9,
             font_color="000000", border=True, font_name=FONTE):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name=font_name, bold=bold, size=font_size, color=font_color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if bg:
        cell.fill = fill(bg)
    if border:
        s = Side(style='thin', color='BFBFBF')
        cell.border = Border(left=s, right=s, top=s, bottom=s)
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def gerar_qdfl(dados: dict, caminho_saida: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "QDFL"

    projeto    = dados.get('projeto', {})
    circuitos  = dados.get('circuitos', [])
    demanda    = dados.get('demanda', {})

    ci = [c for c in circuitos if c.get('potencia_va', 0) > 0]

    # ── Configurar tamanhos das colunas ────────────────────────────
    # Mapeamento fiel ao modelo original
    col_widths = {
        'A':5,   'B':7,   'C':38,
        'D':10,  'E':8,   'F':8,   'G':8,   'H':8,    # TUG slots
        'I':9,   'J':8,   'K':8,   'L':8,   'M':8,    # ILUM slots
        'N':8,   'O':8,   'P':8,   'Q':8,   'R':8,
        'S':8,   'T':8,   'U':8,   'V':8,   'W':8,
        'X':8,   'Y':8,   'Z':10,                       # Carga especial
        'AA':12, 'AB':9,  'AC':13, 'AD':11, 'AE':9,   # Pot, FP, VA, VAr, V
        'AF':10, 'AG':9,  'AH':7,  'AI':11,            # I, Disjuntor N/Curva/Icc
        'AJ':9,  'AK':16, 'AL':10, 'AM':8,  'AN':8,   # DR, Método
        'AO':11, 'AP':8,  'AQ':8,  'AR':8,             # Isolação, Seções
        'AS':8,  'AT':8,  'AU':9,  'AV':10,            # Fa, Ft, Iz nom, Iz real
        'AW':9,  'AX':9,                                # Fase dist, Fase
        'AY':9,  'AZ':9,                                # Fases A/B
        'BA':9,  'BB':9,  'BC':9,                       # V/A.km, dist, dU%
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Alturas
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 15

    # ── LINHA 1: Título ──────────────────────────────────────────
    ws.merge_cells('B1:BC1')
    t = ws['B1']
    t.value     = f"Quadro de Distribuição de Força e Luz — QDFL  |  {projeto.get('nome','')}"
    t.font      = Font(name=FONTE, bold=True, size=12, color=COR_WHITE)
    t.fill      = fill(COR_HEADER_DARK)
    t.alignment = Alignment(horizontal='center', vertical='center')
    s = Side(style='medium', color='000000')
    t.border = Border(left=s, right=s, top=s, bottom=s)

    # Dados do projeto — linha 2
    ws.merge_cells('B2:L2')
    info = ws['B2']
    nom  = projeto.get('nome','')
    end  = projeto.get('endereco','')
    rt   = projeto.get('projetista','')
    crea = projeto.get('crea','')
    vf   = projeto.get('v_fase', 127)
    vl   = projeto.get('v_linha', 220)
    sis  = projeto.get('sistema','Bifasico')
    conc = projeto.get('concessionaria','CEMIG')
    info.value     = f"Obra: {nom} | {end}"
    info.font      = Font(name=FONTE, size=8)
    info.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('M2:Y2')
    ws['M2'].value     = f"RT: {rt} — {crea}"
    ws['M2'].font      = Font(name=FONTE, size=8)
    ws['M2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('Z2:BC2')
    ws['Z2'].value     = f"{conc} | {sis} {vf}/{vl}V | {projeto.get('metodo_instalacao','B1')} | {projeto.get('isolacao','PVC')} {projeto.get('t_amb',30)}°C"
    ws['Z2'].font      = Font(name=FONTE, size=8)
    ws['Z2'].alignment = Alignment(horizontal='right', vertical='center')

    # ── LINHAS 3-5: Cabeçalho ────────────────────────────────────
    def header(ws, r1, r2, col1, col2, text, bg=COR_HEADER_MID):
        if col1 == col2:
            ws.merge_cells(start_row=r1, start_column=col1, end_row=r2, end_column=col2)
        else:
            ws.merge_cells(start_row=r1, start_column=col1, end_row=r2, end_column=col2)
        c = ws.cell(row=r1, column=col1, value=text)
        c.font      = Font(name=FONTE, bold=True, size=8, color=COR_WHITE)
        c.fill      = fill(bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        s = Side(style='thin', color='FFFFFF')
        c.border = Border(left=s, right=s, top=s, bottom=s)

    # Grupos principais (linhas 3-5)
    header(ws, 3, 5,  2,  2, "N°")
    header(ws, 3, 5,  3,  3, "Descrição do Circuito", bg=COR_HEADER_DARK)
    header(ws, 3, 3,  4,  8, "Pontos de Tomadas (W)")
    header(ws, 3, 3,  9, 25, "Pontos de Iluminação (W)")
    header(ws, 3, 5, 26, 26, "Carga Especial (W)")
    header(ws, 3, 5, 27, 27, "Potência Ativa (W)")
    header(ws, 3, 5, 28, 28, "FP")
    header(ws, 3, 5, 29, 29, "Pot. Aparente (VA)")
    header(ws, 3, 5, 30, 30, "Pot. Reativa (VAr)")
    header(ws, 3, 5, 31, 31, "Tensão (V)")
    header(ws, 3, 5, 32, 32, "Corrente (A)")
    header(ws, 3, 4, 33, 35, "Disjuntor")
    header(ws, 3, 4, 36, 37, "Disp. DR")
    header(ws, 3, 4, 38, 44, "Condutor")
    header(ws, 3, 5, 45, 45, "Fa")
    header(ws, 3, 5, 46, 46, "Ft")
    header(ws, 3, 5, 47, 47, "Iz nom. (A)")
    header(ws, 3, 5, 48, 48, "Iz real (A)")
    header(ws, 3, 4, 49, 50, "Balanc. Fases")
    header(ws, 3, 5, 53, 53, "V/A·km")
    header(ws, 3, 5, 54, 54, "dist (km)")
    header(ws, 3, 5, 55, 55, "ΔV%")

    # Sub-cabeçalhos TUG (linha 4-5)
    tug_labels = ['100W','200W','300W','300W','600W']
    for j, lbl in enumerate(tug_labels):
        header(ws, 4, 5, 4+j, 4+j, lbl, bg=COR_HEADER_LIGHT)
        ws.cell(4+j, 4+j).font = Font(name=FONTE, bold=True, size=7, color="000000")

    # Sub-cabeçalhos ILUM (linha 4-5)
    ilum_labels = ['52.5W','60W','60W','100W','100W','150W','150W','150W','150W','150W','150W','150W','200W','250W','300W','300W','500W']
    for j, lbl in enumerate(ilum_labels[:17]):
        header(ws, 4, 5, 9+j, 9+j, lbl, bg=COR_HEADER_LIGHT)
        ws.cell(9+j, 9+j).font = Font(name=FONTE, bold=True, size=7, color="000000")

    # Sub-cabeçalhos Disjuntor
    header(ws, 5, 5, 33, 33, "In (A)")
    header(ws, 5, 5, 34, 34, "Curva")
    header(ws, 5, 5, 35, 35, "Icc (kA)")

    # Sub-cabeçalhos DR
    header(ws, 5, 5, 36, 36, "Sensib.")
    header(ws, 5, 5, 37, 37, "Conjunto")

    # Sub-cabeçalhos Condutor
    header(ws, 5, 5, 38, 38, "Método")
    header(ws, 5, 5, 39, 39, "Classe")
    header(ws, 5, 5, 40, 40, "Isolação")
    header(ws, 5, 5, 41, 41, "Tens.Isol.")
    header(ws, 5, 5, 42, 42, "Fase (mm²)")
    header(ws, 5, 5, 43, 43, "Neutro(mm²)")
    header(ws, 5, 5, 44, 44, "PE (mm²)")

    # Sub-balanc.
    header(ws, 5, 5, 49, 49, "Distrib.")
    header(ws, 5, 5, 50, 50, "Fase")

    # ── LINHAS DE CIRCUITOS ───────────────────────────────────────
    TOTAL_LINHAS = 17  # igual ao modelo (linhas 6 a 22)
    fases_alternas = ['A','B','A','B','A','B','A','B','A','B','A','B','A','B','A','B','A']

    for i in range(TOTAL_LINHAS):
        row = 6 + i
        ws.row_dimensions[row].height = 15
        bg_row = COR_CIRC_ALT if i % 2 == 0 else COR_WHITE

        if i < len(ci):
            c = ci[i]
            desc = c.get('descricao','')
            va      = c.get('potencia_va', 0)
            real_w  = c.get('potencia_real_w', 0) or 0
            tipo = c.get('tipo','')
            fase = c.get('fase','R')
            fp   = 0.95
            pot_ap   = va / fp if fp > 0 else 0
            pot_reat = (pot_ap**2 - va**2)**0.5 if pot_ap > va else 0
            tensao   = c.get('tensao_v', 127)
            ib       = va / tensao if tensao > 0 else 0
            ft_val   = c.get('ft', 1.0)
            fa_val   = c.get('fa', 1.0)
            iz_nom   = c.get('iz_nominal', 0)
            iz_real  = c.get('iz_efetiva', 0)
            secao_f  = c.get('secao_fase', 0)
            secao_n  = c.get('secao_neutro', 0)
            secao_pe = c.get('secao_pe', 0)
            in_disj  = c.get('in_disj', 0)
            curva    = c.get('curva', 'B')
            idr      = c.get('idr', False)
            du       = c.get('du_calc', 0)
            status   = c.get('status', 'SEM_DADOS')
            metodo   = projeto.get('metodo_instalacao','B1')
            isolacao = projeto.get('isolacao','PVC')
            du_disp  = 3.5  # reserva para ramal

            # Cor da linha baseada no status
            if status == 'ERRO':
                bg_row = COR_ERR
            elif status == 'LIMITE':
                bg_row = COR_WARN

            # Colunas de carga — distribuir VA por tipo
            tug_va  = va if tipo == 'TUG' else 0
            ilum_va = va if tipo == 'ILUM' else 0
            esp_va  = va if tipo == 'TUE' else 0

            # Calcular V/A·km e dU
            rho_cu = 0.0172
            alpha  = 0.00393
            rho_t  = rho_cu * (1 + alpha * 50)
            comp   = c.get('comprimento_m', 0)
            n_fases_du = 1 if fase in ['R','S','T'] else 2
            va_km  = (2 * rho_t * 1000) / secao_f if secao_f > 0 else 0

            # Fase distribuição
            fase_dist = fases_alternas[i]

            # Preencher célula a célula
            def sc(col, val, **kw):
                kw.setdefault('bg', bg_row)
                set_cell(ws, row, col, val, **kw)

            sc(2,  i+1,          halign='center', bold=True)
            # Indicar pot. real no nome quando for diferente (ex: LED 54W)
            desc_exib = desc
            if real_w and real_w > 0 and abs(real_w - va) > 10:
                desc_exib = f"{desc} [{real_w}W real]"
            sc(3,  desc_exib,    halign='left', wrap=False)

            # Cargas TUG (colunas 4-8) e ILUM (9-25)
            if tipo == 'TUG' and tug_va > 0:
                # Distribuir nas colunas de tomada mais adequadas
                sc(4, 0); sc(5, 0); sc(6, 0); sc(7, 0); sc(8, 0)
                # Usar col D para tomadas de 100W count
                n_tom = max(1, round(tug_va / 100))
                sc(4, n_tom, num_fmt='0')
            elif tipo == 'ILUM' and ilum_va > 0:
                for col_i in range(4, 26):
                    sc(col_i, 0)
                sc(9, round(ilum_va / 52.5, 1) if ilum_va > 0 else 0, num_fmt='0.0')

            sc(26, esp_va if tipo == 'TUE' else 0, num_fmt='#,##0')
            sc(27, va,          halign='right', num_fmt='#,##0')
            sc(28, fp,          halign='center', num_fmt='0.00')
            sc(29, pot_ap,      halign='right',  num_fmt='#,##0.00')
            sc(30, pot_reat,    halign='right',  num_fmt='#,##0.00')
            sc(31, tensao,      halign='center', num_fmt='0')
            sc(32, ib,          halign='right',  num_fmt='0.00')

            # Disjuntor
            sc(33, in_disj,     halign='center', num_fmt='0')
            sc(34, curva,       halign='center')
            sc(35, f"{projeto.get('icc_rede_ka',3):.0f}kA", halign='center')

            # DR
            if idr:
                sc(36, '25mA',  halign='center', bg='FFE699')
                sc(37, 'Conj.1',halign='center', bg='FFE699')
            else:
                sc(36, '—',     halign='center')
                sc(37, '—',     halign='center')

            # Condutor
            sc(38, metodo,      halign='center')
            sc(39, '5',         halign='center')
            sc(40, isolacao,    halign='center')
            sc(41, '750V',      halign='center')
            sc(42, secao_f,     halign='center', num_fmt='0.0',
               bg=COR_OK if secao_f > 0 else bg_row)
            sc(43, secao_n if secao_n else '—', halign='center')
            sc(44, secao_pe,    halign='center', num_fmt='0.0')

            # Fatores
            sc(45, fa_val,      halign='center', num_fmt='0.000')
            sc(46, ft_val,      halign='center', num_fmt='0.000')
            sc(47, iz_nom,      halign='right',  num_fmt='0.0')
            sc(48, iz_real,     halign='right',  num_fmt='0.0',
               bg=COR_OK if iz_real > ib else COR_ERR)

            # Balanceamento
            sc(49, fase_dist,   halign='center')
            sc(50, fase,        halign='center')

            # Queda de tensão
            sc(53, round(va_km,1), halign='right', num_fmt='0.0')
            sc(54, comp/1000,   halign='right',  num_fmt='0.000')
            du_bg = COR_OK if du <= 3.5 else (COR_WARN if du <= 4 else COR_ERR)
            sc(55, du,          halign='right',  num_fmt='0.00', bg=du_bg)

        else:
            # Linha vazia com estrutura
            for col in range(2, 56):
                set_cell(ws, row, col, None, bg=bg_row)
            # Fase alternada
            set_cell(ws, row, 49, fases_alternas[i], bg=bg_row, halign='center')
            set_cell(ws, row, 50, '—', bg=bg_row, halign='center')

    # ── LINHA DE TOTAIS ───────────────────────────────────────────
    row_total = 6 + TOTAL_LINHAS
    ws.row_dimensions[row_total].height = 18
    ws.merge_cells(f'B{row_total}:C{row_total}')

    def sc_tot(col, val, **kw):
        kw.setdefault('bg', COR_TOTAL)
        kw.setdefault('bold', True)
        set_cell(ws, row_total, col, val, **kw)

    # Calcular totais
    total_va = sum(c.get('potencia_va',0) for c in ci)
    fp_medio = 0.95
    pot_ap_total = total_va / fp_medio if fp_medio > 0 else 0
    i_dem        = demanda.get('i_dem', 0)
    in_geral     = demanda.get('in_geral', 0)
    dem_kw       = demanda.get('dem_kw', 0)
    ci_kw        = demanda.get('ci_kw', 0)
    v_linha      = projeto.get('v_linha', 220)
    metodo       = projeto.get('metodo_instalacao','B1')
    isolacao     = projeto.get('isolacao','PVC')

    ws.cell(row=row_total, column=2).value = 'QDFL'
    ws.cell(row=row_total, column=2).font  = Font(name=FONTE, bold=True, size=9, color=COR_WHITE)
    ws.cell(row=row_total, column=2).fill  = fill(COR_HEADER_DARK)
    ws.cell(row=row_total, column=2).alignment = Alignment(horizontal='center', vertical='center')

    sc_tot(27, total_va,      num_fmt='#,##0',    halign='right')
    sc_tot(28, fp_medio,      num_fmt='0.00',      halign='center')
    sc_tot(29, pot_ap_total,  num_fmt='#,##0.00',  halign='right')
    sc_tot(31, v_linha,       num_fmt='0',          halign='center')
    sc_tot(32, i_dem,         num_fmt='0.00',       halign='right')
    sc_tot(33, in_geral,      num_fmt='0',          halign='center',
           bg=COR_OK if in_geral > 0 else COR_TOTAL)
    sc_tot(38, metodo,        halign='center')
    sc_tot(39, '5',           halign='center')
    sc_tot(40, isolacao,      halign='center')
    sc_tot(41, '0,6/1kV',     halign='center')

    # Seções do ramal
    ramal = demanda.get('ramal_min_mm2', 10)
    sc_tot(42, ramal, num_fmt='0', halign='center')
    sc_tot(43, ramal, num_fmt='0', halign='center')
    sc_tot(44, ramal, num_fmt='0', halign='center')
    sc_tot(45, 1.0,   num_fmt='0.000', halign='center')
    sc_tot(46, 1.0,   num_fmt='0.000', halign='center')
    sc_tot(49, 'ABC', halign='center')

    # ── LINHA DE RESUMO DEMANDA ───────────────────────────────────
    row_res = row_total + 2
    ws.row_dimensions[row_res].height = 14
    ws.row_dimensions[row_res + 1].height = 14
    ws.row_dimensions[row_res + 2].height = 14

    resumo = [
        ('CI instalada', f"{ci_kw:.3f} kW"),
        ('FD (CEMIG ND-5.1)', f"{demanda.get('fd',1)*100:.0f}%"),
        ('Demanda máxima', f"{dem_kw:.3f} kW"),
        ('Tipo ligação CEMIG', demanda.get('tipo_ligacao_cemig','')),
        ('QD: posições', f"{demanda.get('n_total_qd',0)} ({demanda.get('n_ativos',0)}+{demanda.get('n_reservas',0)} res.)"),
    ]
    for j, (k, v) in enumerate(resumo):
        ws.merge_cells(f'B{row_res+j}:J{row_res+j}')
        ws.cell(row_res+j, 2).value = k
        ws.cell(row_res+j, 2).font  = Font(name=FONTE, bold=True, size=8)
        ws.cell(row_res+j, 2).alignment = Alignment(horizontal='left')
        ws.merge_cells(f'K{row_res+j}:R{row_res+j}')
        ws.cell(row_res+j, 11).value = v
        ws.cell(row_res+j, 11).font  = Font(name=FONTE, size=8)
        ws.cell(row_res+j, 11).alignment = Alignment(horizontal='left')

    # ── ABA DE REFERÊNCIA (Tabelas NBR 5410) ────────────────────
    ws_ref = wb.create_sheet("Ref. NBR 5410")
    ref_headers = [
        "Tabela 36 — Capacidade de Condução (Iz) por Método de Instalação",
        "PVC 70°C — Cobre — Método B1 (eletroduto embutido em alvenaria)",
        "",
        "Seção (mm²)", "2 Cond. (A)", "3 Cond. (A)",
    ]
    for i, h in enumerate(ref_headers[:3]):
        ws_ref.merge_cells(f'A{i+1}:F{i+1}')
        c = ws_ref.cell(i+1, 1, h)
        c.font      = Font(name=FONTE, bold=True, size=9, color=COR_WHITE if i < 2 else "000000")
        c.fill      = fill(COR_HEADER_DARK if i == 0 else COR_HEADER_MID if i == 1 else COR_WHITE)
        c.alignment = Alignment(horizontal='center')

    tabela_iz = [
        (1.5, 17.5, 15.5), (2.5, 24, 21), (4, 32, 28), (6, 41, 36),
        (10, 57, 50), (16, 76, 68), (25, 101, 89), (35, 125, 110),
        (50, 151, 134), (70, 192, 171), (95, 232, 207), (120, 269, 239),
    ]
    for j, (s, iz2, iz3) in enumerate(tabela_iz):
        r = 4 + j
        bg = COR_CIRC_ALT if j % 2 == 0 else COR_WHITE
        for col, v in [(1, s), (2, iz2), (3, iz3)]:
            c = ws_ref.cell(r, col, v)
            c.font      = Font(name=FONTE, size=9)
            c.alignment = Alignment(horizontal='center')
            c.fill      = fill(bg)

    wb.save(caminho_saida)
    print(f"QDFL salvo em: {caminho_saida}")

# ── Entry point ─────────────────────────────────────────────────
if __name__ == '__main__':
    if len(sys.argv) < 3:
        # Teste com dados de exemplo
        dados_teste = {
            "projeto": {
                "nome": "Residência Araguari",
                "endereco": "Rua das Flores, 123 — Araguari/MG",
                "projetista": "Wiliam A. S. Junior",
                "crea": "CREA-MG 23564875",
                "sistema": "Bifasico",
                "v_fase": 127, "v_linha": 220,
                "metodo_instalacao": "B1",
                "isolacao": "PVC",
                "t_amb": 30, "icc_rede_ka": 3,
                "concessionaria": "CEMIG",
            },
            "circuitos": [
                {"descricao":"ILUM: Sala de Estar","potencia_va":315,"tipo":"ILUM","fase":"R",
                 "comprimento_m":18,"ft":1.0,"fa":1.0,"iz_nominal":17.5,"iz_efetiva":17.5,
                 "secao_fase":1.5,"secao_neutro":1.5,"secao_pe":1.5,"in_disj":10,
                 "curva":"B","idr":False,"du_calc":1.59,"tensao_v":127,"ib":2.48,
                 "status":"OK","violacoes":[]},
                {"descricao":"TUG: Sala de Estar","potencia_va":1000,"tipo":"TUG","fase":"S",
                 "comprimento_m":12,"ft":1.0,"fa":1.0,"iz_nominal":24.0,"iz_efetiva":24.0,
                 "secao_fase":2.5,"secao_neutro":2.5,"secao_pe":2.5,"in_disj":16,
                 "curva":"B","idr":False,"du_calc":2.33,"tensao_v":127,"ib":7.87,
                 "status":"OK","violacoes":[]},
                {"descricao":"TUE: Chuveiro (Banheiro)","potencia_va":5500,"tipo":"TUE","fase":"RS",
                 "comprimento_m":15,"ft":1.0,"fa":1.0,"iz_nominal":32.0,"iz_efetiva":32.0,
                 "secao_fase":4,"secao_neutro":4,"secao_pe":4,"in_disj":25,
                 "curva":"C","idr":True,"du_calc":1.52,"tensao_v":220,"ib":25.0,
                 "status":"OK","violacoes":[]},
            ],
            "demanda": {
                "ci_kw": 6.815, "fd": 0.87, "dem_kw": 5.929,
                "i_dem": 15.6, "in_geral": 20,
                "tipo_ligacao_cemig": "B1",
                "ramal_min_mm2": 10,
                "n_ativos": 3, "n_reservas": 1, "n_total_qd": 4,
            }
        }
        gerar_qdfl(dados_teste, '/mnt/user-data/outputs/QDFL_teste.xlsx')
    else:
        with open(sys.argv[1]) as f:
            dados = json.load(f)
        gerar_qdfl(dados, sys.argv[2])
